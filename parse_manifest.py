#!/usr/bin/env python3
"""
Parse a container manifest CSV and update inventory_master.xlsx.

For each manifest line item:
  MATCHED variant  → find In Production rows (blank serial), fill serials,
                     flip status to "Pre-Sale - <container_id>"
  NEW variant      → insert a new block at the end of that design's section
                     (or at sheet end for a brand-new design) with Pre-Sale rows

Spacing rules preserved:
  • 1 blank row  between variants of the same design
  • 2 blank rows between different designs

Usage:
    python3 parse_manifest.py --container CNT-001
    python3 parse_manifest.py --container CNT-001 --manifest container_manifest.csv
"""

import argparse
import csv

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from apply_status_fills import apply_fills
from fix_inventory_formatting import apply_design_block_borders
from refresh_counts import refresh_counts

INVENTORY_FILE = "inventory_master.xlsx"
MANIFEST_FILE  = "container_manifest.csv"

DATA_START   = 3
VARIANT_KEYS = ["Design Name", "Size", "Finish", "Swing", "Glass Type"]
C_SERIAL     = 11   # K
C_STATUS     = 12   # L

WHITE = "FFFFFFFF"
BLACK = "FF000000"


# ── Helpers ───────────────────────────────────────────────────────────────────

def _font(bold=False, color=BLACK):
    return Font(name="Arial", size=10, bold=bold, color=color)

def _center():
    return Alignment(horizontal="center")

def _left():
    return Alignment(horizontal="left")


# ── Manifest ──────────────────────────────────────────────────────────────────

def load_manifest(path):
    with open(path, newline="") as f:
        return list(csv.DictReader(f))

def manifest_key(row):
    return tuple(row[k].strip() for k in VARIANT_KEYS)

def manifest_serials(row):
    return [s.strip() for s in row["Serial Numbers"].split(";") if s.strip()]


# ── Inventory scan ────────────────────────────────────────────────────────────

def scan_inventory(ws):
    """
    Returns:
      variant_rows : {variant_key: [row numbers]}
      design_last  : {design_name: last data row}   last row with content for that design
      design_order : [design_name, ...]              insertion order of designs
    """
    variant_rows = {}
    design_last  = {}
    design_order = []

    for r in range(DATA_START, ws.max_row + 1):
        key = tuple(str(ws.cell(r, c).value or "").strip() for c in range(1, 6))
        if not any(key):
            continue
        if key not in variant_rows:
            variant_rows[key] = []
        variant_rows[key].append(r)

        design = key[0]
        if design not in design_last:
            design_order.append(design)
        design_last[design] = r

    return variant_rows, design_last, design_order


def in_production_rows(ws, rows):
    return [r for r in rows
            if ws.cell(r, C_STATUS).value == "In Production"
            and not ws.cell(r, C_SERIAL).value]


# ── Write helpers ─────────────────────────────────────────────────────────────

def write_detail_row(ws, r, variant, serial, status, first=False):
    for c, val in enumerate(variant, start=1):
        cell = ws.cell(r, c)
        cell.value     = val
        cell.font      = _font(bold=True) if first else _font(color=WHITE)
        cell.alignment = _left()
    ws.cell(r, C_SERIAL).value     = serial
    ws.cell(r, C_SERIAL).font      = _font()
    ws.cell(r, C_SERIAL).alignment = _center()
    ws.cell(r, C_STATUS).value     = status
    ws.cell(r, C_STATUS).font      = _font()
    ws.cell(r, C_STATUS).alignment = _center()


# ── Core logic ────────────────────────────────────────────────────────────────

def build_insertion_plan(new_items, variant_rows, design_last, design_order, last_sheet_row):
    """
    Group new variants by design.  For each design build one contiguous block
    of rows to insert, then return a sorted list of:
        (insert_after_row, rows_to_write)

    New designs are inserted in alphabetical order relative to existing designs.
    Sorted descending so callers execute bottom-to-top without shifting earlier positions.
    """
    by_design = {}
    for key, serials, status in new_items:
        by_design.setdefault(key[0], []).append((key, serials, status))

    # For brand-new designs, find the correct alphabetical insertion point.
    # That is: insert after the last row of the last existing design whose
    # name sorts before the new design name.
    def insert_after_for_new_design(new_design):
        preceding = [d for d in design_order if d < new_design]
        if preceding:
            return design_last[max(preceding)]
        # New design sorts before all existing — insert right after header/blank rows
        return DATA_START - 1

    plan = []

    for design, items in by_design.items():
        if design in design_last:
            insert_after   = design_last[design]
            leading_blanks = 1   # separator between variants of same design
        else:
            insert_after   = insert_after_for_new_design(design)
            leading_blanks = 2   # separator between designs

        rows_to_write = []
        for idx, (key, serials, status) in enumerate(items):
            n_blanks = leading_blanks if idx == 0 else 1
            rows_to_write.extend([None] * n_blanks)
            for serial in serials:
                rows_to_write.append((key, serial, status))

        plan.append((insert_after, rows_to_write))

    plan.sort(key=lambda x: x[0], reverse=True)
    return plan


def execute_plan(ws, plan):
    for insert_after, rows_to_write in plan:
        n = len(rows_to_write)
        ws.insert_rows(insert_after + 1, amount=n)
        prev_blank = True
        for i, row_spec in enumerate(rows_to_write):
            r = insert_after + 1 + i
            if row_spec is None:
                for c in range(1, C_STATUS + 1):
                    ws.cell(r, c).value = None
                prev_blank = True
            else:
                key, serial, status = row_spec
                write_detail_row(ws, r, key, serial, status, first=prev_blank)
                prev_blank = False


def process_manifest(ws, manifest_rows, container_id):
    pre_sale_status = f"Pre-Sale - {container_id}"
    variant_rows, design_last, design_order = scan_inventory(ws)

    matched_items = []
    new_items     = []

    for item in manifest_rows:
        key     = manifest_key(item)
        serials = manifest_serials(item)
        if key in variant_rows:
            matched_items.append((key, serials))
        else:
            new_items.append((key, serials, pre_sale_status))

    # ── Matched: update In Production rows in place ────────────────────────────
    for key, serials in matched_items:
        prod_rows = in_production_rows(ws, variant_rows[key])
        qty = len(serials)

        if len(prod_rows) != qty:
            print(f"  WARNING: {key[0]} {key[1]} — manifest has {qty} serial(s) "
                  f"but {len(prod_rows)} In Production row(s). "
                  f"Filling {min(len(prod_rows), qty)}.")

        for i in range(min(len(prod_rows), qty)):
            r = prod_rows[i]
            ws.cell(r, C_SERIAL).value     = serials[i]
            ws.cell(r, C_SERIAL).font      = _font()
            ws.cell(r, C_SERIAL).alignment = _center()
            ws.cell(r, C_STATUS).value     = pre_sale_status
            ws.cell(r, C_STATUS).font      = _font()
            ws.cell(r, C_STATUS).alignment = _center()

        filled = min(len(prod_rows), qty)
        if filled:
            print(f"  MATCHED   {key[0]} {key[1]} {key[2]} — "
                  f"{filled} serial(s) → {pre_sale_status}")

    # ── New: build plan and insert bottom-to-top ───────────────────────────────
    if new_items:
        plan = build_insertion_plan(
            new_items, variant_rows, design_last, design_order, ws.max_row
        )
        execute_plan(ws, plan)

        for key, serials, _ in new_items:
            tag = "NEW DESIGN" if key[0] not in design_last else "NEW VARIANT"
            print(f"  {tag:<12}{key[0]} {key[1]} {key[2]} — "
                  f"{len(serials)} row(s) → {pre_sale_status}")

    return len(matched_items), len(new_items)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Parse a container manifest and update inventory status."
    )
    parser.add_argument("--container", required=True, help="Container ID, e.g. CNT-001")
    parser.add_argument("--manifest",  default=MANIFEST_FILE)
    parser.add_argument("--inventory", default=INVENTORY_FILE)
    args = parser.parse_args()

    print(f"Loading manifest: {args.manifest}")
    manifest_rows = load_manifest(args.manifest)
    print(f"  {len(manifest_rows)} line item(s)\n")

    wb = load_workbook(args.inventory)
    ws = wb.active

    n_matched, n_new = process_manifest(ws, manifest_rows, args.container)

    print(f"\nApplying formatting...")
    apply_design_block_borders(ws)
    refresh_counts(ws)
    apply_fills(ws)

    wb.save(args.inventory)
    print(f"Saved → {args.inventory}")
    print(f"\nSummary: {n_matched} matched, {n_new} new variant(s) created.")


if __name__ == "__main__":
    main()
