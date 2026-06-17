#!/usr/bin/env python3
"""
Build inventory_master.xlsx from scratch with sample steel door product data.

Layout
------
Row 1  : Container group labels (CNT-001 / CNT-002 / CNT-003 over cols K-M)
Row 2  : Column headers
Row 3+ : One summary row per product variant, followed by SERIAL_ROWS blank
         rows reserved for serial inserts (done by inventory_update.py).

Compatibility note
------------------
inventory_update.py reads its header map from ws[1].  With this two-row
header layout it will need one change: replace `ws[1]` with `ws[2]` and
update VARIANT_KEYS / column-name constants to match the names used here.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# ── Column indices (1-based) ──────────────────────────────────────────────────
C_DESIGN   = 1   # A  Design Name
C_SIZE     = 2   # B  Size
C_FINISH   = 3   # C  Finish
C_SWING    = 4   # D  Swing
C_GLASS    = 5   # E  Glass Type
C_INSTOCK  = 6   # F  In-Stock QTY
C_INPROD   = 7   # G  In-Production QTY
C_ONCONT   = 8   # H  On-Container QTY
C_OPTIMAL  = 9   # I  Optimal Count
C_VARIANCE = 10  # J  Variance  (live formula)
C_CNT001   = 11  # K  CNT-001 → Presale QTY
C_CNT002   = 12  # L  CNT-002 → Presale QTY
C_CNT003   = 13  # M  CNT-003 → Presale QTY

TOTAL_COLS  = C_CNT003
GROUP_ROW   = 1   # container-name labels
HEADER_ROW  = 2   # column headers
DATA_START  = 3   # first product summary row
SERIAL_ROWS = 5   # blank rows reserved after each summary row for serials

# ── Sample product data ───────────────────────────────────────────────────────
# (design, size, finish, swing, glass,
#  in_stock, in_prod, on_cont, optimal,
#  cnt001_presale, cnt002_presale, cnt003_presale)
PRODUCTS = [
    # Valencia – 3 variants
    ("Valencia", '32" x 80"', "Matte Black",       "Left Inswing",   "Clear",    4, 2, 1, 6,  1, 0, 0),
    ("Valencia", '36" x 80"', "Matte Black",       "Right Inswing",  "Clear",    7, 0, 3, 8,  2, 1, 0),
    ("Valencia", '36" x 96"', "Oil-Rubbed Bronze", "Right Outswing", "Frosted",  2, 4, 0, 6,  0, 2, 0),
    # Sevilla – 2 variants
    ("Sevilla",  '32" x 80"', "Brushed Nickel",    "Right Inswing",  "Rain",     5, 1, 2, 7,  1, 0, 1),
    ("Sevilla",  '36" x 80"', "Satin White",       "Left Inswing",   "None",     3, 3, 0, 8,  0, 1, 0),
    # Cordova – 3 variants
    ("Cordova",  '32" x 80"', "Matte Black",       "Right Inswing",  "Frosted",  6, 0, 2, 7,  2, 0, 0),
    ("Cordova",  '36" x 80"', "Oil-Rubbed Bronze", "Left Inswing",   "Clear",    3, 2, 3, 9,  1, 1, 1),
    ("Cordova",  '36" x 96"', "Brushed Nickel",    "Right Outswing", "Rain",     1, 5, 0, 7,  0, 2, 0),
    # Granada – 2 variants
    ("Granada",  '36" x 80"', "Matte Black",       "Right Inswing",  "Clear",    8, 0, 1, 8,  1, 0, 0),
    ("Granada",  '36" x 96"', "Satin White",       "Left Inswing",   "Sidelite", 2, 3, 2, 6,  0, 1, 0),
    # Malaga – 2 variants
    ("Malaga",   '32" x 80"', "Brushed Nickel",    "Left Inswing",   "None",     4, 1, 1, 5,  0, 0, 1),
    ("Malaga",   '36" x 80"', "Oil-Rubbed Bronze", "Right Inswing",  "Frosted",  0, 2, 4, 8,  1, 2, 0),
    # Altea – 3 variants
    ("Altea",    '32" x 80"', "Satin White",       "Right Inswing",  "Clear",    5, 2, 0, 6,  1, 0, 0),
    ("Altea",    '36" x 80"', "Matte Black",       "Left Inswing",   "Rain",     3, 0, 3, 7,  0, 1, 1),
    ("Altea",    '36" x 96"', "Oil-Rubbed Bronze", "Right Outswing", "Sidelite", 1, 4, 2, 8,  0, 0, 2),
]

# ── Container groups ──────────────────────────────────────────────────────────
CONTAINERS = [
    ("CNT-001", C_CNT001),
    ("CNT-002", C_CNT002),
    ("CNT-003", C_CNT003),
]

# ── Column display widths ─────────────────────────────────────────────────────
COL_WIDTHS = {
    C_DESIGN:   18,
    C_SIZE:     12,
    C_FINISH:   21,
    C_SWING:    16,
    C_GLASS:    13,
    C_INSTOCK:  14,
    C_INPROD:   20,
    C_ONCONT:   20,
    C_OPTIMAL:  14,
    C_VARIANCE: 12,
    C_CNT001:   16,
    C_CNT002:   16,
    C_CNT003:   16,
}


# ── Style helpers ─────────────────────────────────────────────────────────────
def thin_border():
    s = Side(border_style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def make_font(bold=False, color="000000", size=10):
    return Font(name="Arial", size=size, bold=bold, color=color)


def solid_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center")

BORDER = thin_border()

NUMERIC_COLS = {C_INSTOCK, C_INPROD, C_ONCONT, C_OPTIMAL, C_VARIANCE,
                C_CNT001, C_CNT002, C_CNT003}


# ── Build workbook ────────────────────────────────────────────────────────────
def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Master"

    # ── Row 1: container group labels ─────────────────────────────────────────
    for col in range(1, TOTAL_COLS + 1):
        cell = ws.cell(row=GROUP_ROW, column=col)
        cell.fill   = solid_fill("D9E2F3")
        cell.border = BORDER

    for label, col in CONTAINERS:
        cell = ws.cell(row=GROUP_ROW, column=col, value=label)
        cell.font      = make_font(bold=True, color="1F3864", size=10)
        cell.fill      = solid_fill("C5D9F1")
        cell.alignment = ALIGN_CENTER
        cell.border    = BORDER

    ws.row_dimensions[GROUP_ROW].height = 18

    # ── Row 2: column headers ─────────────────────────────────────────────────
    column_headers = [
        "Design Name", "Size", "Finish", "Swing", "Glass Type",
        "In-Stock QTY", "In-Production QTY", "On-Container QTY",
        "Optimal Count", "Variance",
        "Presale QTY", "Presale QTY", "Presale QTY",
    ]

    for col, header in enumerate(column_headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col, value=header)
        cell.font      = make_font(bold=True)
        cell.fill      = solid_fill("BDD7EE")
        cell.alignment = ALIGN_CENTER
        cell.border    = BORDER

    ws.row_dimensions[HEADER_ROW].height = 20

    # ── Product summary rows ──────────────────────────────────────────────────
    last_data_row = DATA_START  # updated as we go

    for i, product in enumerate(PRODUCTS):
        design, size, finish, swing, glass, \
            in_stock, in_prod, on_cont, optimal, \
            cnt001, cnt002, cnt003 = product

        row = DATA_START + i * (1 + SERIAL_ROWS)
        last_data_row = row

        f = get_column_letter(C_INSTOCK)
        g = get_column_letter(C_INPROD)
        h = get_column_letter(C_ONCONT)
        opt = get_column_letter(C_OPTIMAL)
        variance_formula = f"={f}{row}+{g}{row}+{h}{row}-{opt}{row}"

        row_values = [
            design, size, finish, swing, glass,
            in_stock, in_prod, on_cont, optimal, variance_formula,
            cnt001, cnt002, cnt003,
        ]

        for col, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font      = make_font()
            cell.border    = BORDER
            cell.alignment = ALIGN_CENTER if col in NUMERIC_COLS else ALIGN_LEFT

        ws.row_dimensions[row].height = 16

    # ── Conditional formatting: Variance < 0 → red fill + red font ───────────
    variance_col   = get_column_letter(C_VARIANCE)
    variance_range = f"{variance_col}{DATA_START}:{variance_col}{last_data_row}"

    ws.conditional_formatting.add(
        variance_range,
        CellIsRule(
            operator="lessThan",
            formula=["0"],
            fill=solid_fill("FFCCCC"),
            font=Font(name="Arial", size=10, color="C00000", bold=True),
        ),
    )

    # ── Column widths ─────────────────────────────────────────────────────────
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Freeze panes: rows 1-2 pinned, col A pinned ───────────────────────────
    ws.freeze_panes = ws.cell(row=DATA_START, column=C_SIZE)

    # ── Sheet-level cosmetics ─────────────────────────────────────────────────
    ws.sheet_view.showGridLines = True

    wb.save("inventory_master.xlsx")
    print(f"Saved inventory_master.xlsx")
    print(f"  {len(PRODUCTS)} product rows, {SERIAL_ROWS} blank rows each for serial inserts")
    print(f"  Data range: rows {DATA_START}–{last_data_row}  |  Columns A–{get_column_letter(TOTAL_COLS)}")
    print(f"  Variance conditional formatting: {variance_range}")


if __name__ == "__main__":
    main()
