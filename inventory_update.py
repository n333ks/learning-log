import csv
from openpyxl import load_workbook

# ── CONFIG ───────────────────────────────────────────
MANIFEST_FILE = "container_manifest.csv"
INVENTORY_FILE = "inventory_master.xlsx"

# Columns that together identify a unique product variant.
# Must match the column headers in both the manifest CSV and the inventory sheet.
VARIANT_KEYS = ["Design Name", "Size", "Finish", "Glass Type", "Top Shape", "Hardware", "Swing"]

PRESALE_COL = "Presale Qty"
IN_STOCK_COL = "In-Stock Qty"
SERIAL_COL = "Serial Number"


# ── LOAD MANIFEST ────────────────────────────────────
def load_manifest(filename):
    with open(filename, newline="") as f:
        reader = csv.DictReader(f)
        return list(reader)


def variant_key(row, keys=VARIANT_KEYS):
    return tuple(row.get(k, "").strip() for k in keys)


# ── LOCATE PRODUCT ROWS IN INVENTORY SHEET ───────────
def find_product_rows(ws, header_to_col):
    """Map each existing product's variant key to its sheet row number."""
    rows_by_variant = {}
    for row_num in range(2, ws.max_row + 1):
        key = tuple(
            str(ws.cell(row=row_num, column=header_to_col[k]).value or "").strip()
            for k in VARIANT_KEYS
        )
        if any(key):
            rows_by_variant[key] = row_num
    return rows_by_variant


# ── MAIN UPDATE LOGIC ─────────────────────────────────
def update_inventory(manifest_items, ws):
    header_to_col = {cell.value: cell.column for cell in ws[1]}
    for required in VARIANT_KEYS + [PRESALE_COL, IN_STOCK_COL, SERIAL_COL]:
        if required not in header_to_col:
            raise ValueError(f"Inventory sheet is missing expected column: {required}")

    rows_by_variant = find_product_rows(ws, header_to_col)

    # Resolve manifest line items to existing sheet rows up front, then apply
    # updates bottom-to-top so inserting serial rows never shifts a row we
    # still need to process.
    matches = []
    unmatched = []
    for item in manifest_items:
        key = variant_key(item)
        if key in rows_by_variant:
            matches.append((rows_by_variant[key], item))
        else:
            unmatched.append(item)

    matches.sort(key=lambda m: m[0], reverse=True)

    for row_num, item in matches:
        qty = int(item["Quantity"])
        presale_cell = ws.cell(row=row_num, column=header_to_col[PRESALE_COL])
        in_stock_cell = ws.cell(row=row_num, column=header_to_col[IN_STOCK_COL])

        presale_qty = int(presale_cell.value or 0)
        in_stock_qty = int(in_stock_cell.value or 0)

        if qty > presale_qty:
            print(
                f"  WARNING: {item['Design Name']} ({item['Size']}, {item['Swing']}) "
                f"moving {qty} but only {presale_qty} in presale."
            )

        presale_cell.value = presale_qty - qty
        in_stock_cell.value = in_stock_qty + qty

        serials = [s for s in item.get("Serial Numbers", "").split(";") if s]
        if serials:
            ws.insert_rows(row_num + 1, amount=len(serials))
            for offset, serial in enumerate(serials):
                ws.cell(
                    row=row_num + 1 + offset,
                    column=header_to_col[SERIAL_COL],
                    value=serial,
                )

        print(f"  Updated {item['Design Name']} ({item['Size']}, {item['Swing']}): "
              f"presale {presale_qty} -> {presale_qty - qty}, "
              f"in-stock {in_stock_qty} -> {in_stock_qty + qty}, "
              f"{len(serials)} serial(s) added")

    return unmatched


# ── MAIN ─────────────────────────────────────────────
def main():
    print("Loading manifest...")
    manifest_items = load_manifest(MANIFEST_FILE)
    print(f"Found {len(manifest_items)} line items.\n")

    print("Opening inventory workbook...")
    wb = load_workbook(INVENTORY_FILE)
    ws = wb.active

    print("Updating inventory...")
    unmatched = update_inventory(manifest_items, ws)

    wb.save(INVENTORY_FILE)
    print(f"\nSaved updates to {INVENTORY_FILE}")

    if unmatched:
        print(f"\n{len(unmatched)} manifest line item(s) had no matching inventory row:")
        for item in unmatched:
            print(f"  - {item['Design Name']} | {item['Size']} | {item['Finish']} | "
                  f"{item['Glass Type']} | {item['Top Shape']} | {item['Hardware']} | {item['Swing']}")


if __name__ == "__main__":
    main()
