#!/usr/bin/env python3
"""
Sort inventory_master.xlsx design blocks alphabetically by Design Name.
Variant order within each design is preserved.
Spacing rules: 1 blank between variants of same design, 2 blanks between designs.
"""

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from apply_status_fills import apply_fills
from fix_inventory_formatting import apply_design_block_borders
from refresh_counts import refresh_counts

INVENTORY_FILE = "inventory_master.xlsx"
DATA_START = 3
C_SERIAL   = 11
C_STATUS   = 12
C_OPTIMAL  = 9

WHITE = "FFFFFFFF"
BLACK = "FF000000"


def _font(bold=False, color=BLACK):
    return Font(name="Arial", size=10, bold=bold, color=color)


def extract_blocks(ws):
    """
    Returns an ordered list of design blocks:
      [
        {
          'design': str,
          'variants': [
            {'key': tuple, 'optimal': int, 'rows': [(serial, status), ...]},
            ...
          ]
        },
        ...
      ]
    Preserves original variant order within each design.
    """
    design_blocks = {}   # design_name -> {variants ordered dict}
    design_order  = []   # preserve insertion order for within-design variant seq

    for r in range(DATA_START, ws.max_row + 1):
        key = tuple(str(ws.cell(r, c).value or "").strip() for c in range(1, 6))
        if not any(key):
            continue

        design = key[0]
        if design not in design_blocks:
            design_blocks[design] = {}
            design_order.append(design)

        if key not in design_blocks[design]:
            optimal = ws.cell(r, C_OPTIMAL).value
            design_blocks[design][key] = {"key": key, "optimal": optimal, "rows": []}

        serial = ws.cell(r, C_SERIAL).value
        status = ws.cell(r, C_STATUS).value
        if serial or status:
            design_blocks[design][key]["rows"].append((serial, status))

    # Convert to list form, keeping variant insertion order within each design
    result = []
    for design in design_order:
        variants = list(design_blocks[design].values())
        result.append({"design": design, "variants": variants})

    return result


def write_blocks(ws, blocks):
    """Clear all data rows and rewrite blocks in the given order."""
    # Clear everything from DATA_START down
    for r in range(DATA_START, ws.max_row + 1):
        for c in range(1, C_STATUS + 1):
            ws.cell(r, c).value = None

    row = DATA_START

    for b_idx, block in enumerate(blocks):
        # Two blank rows before each design (except the very first)
        if b_idx > 0:
            row += 2

        for v_idx, variant in enumerate(block["variants"]):
            # One blank row between variants of same design
            if v_idx > 0:
                row += 1

            key = variant["key"]
            left_align = Alignment(horizontal="left")
            center     = Alignment(horizontal="center")

            for i, (serial, status) in enumerate(variant["rows"]):
                # First row of each variant: bold black (visible summary)
                # Subsequent rows: white (hidden repeated labels)
                a_e_font = _font(bold=(i == 0)) if i == 0 else _font(color=WHITE)
                for c, val in enumerate(key, start=1):
                    cell = ws.cell(row, c)
                    cell.value     = val
                    cell.font      = a_e_font
                    cell.alignment = left_align
                ws.cell(row, C_SERIAL).value     = serial
                ws.cell(row, C_SERIAL).font      = _font()
                ws.cell(row, C_SERIAL).alignment = center
                ws.cell(row, C_STATUS).value     = status
                ws.cell(row, C_STATUS).font      = _font()
                ws.cell(row, C_STATUS).alignment = center
                row += 1

    return row - 1   # last written row


def main():
    wb = load_workbook(INVENTORY_FILE)
    ws = wb.active

    blocks = extract_blocks(ws)

    # Sort blocks alphabetically by design name
    blocks.sort(key=lambda b: b["design"])

    last_row = write_blocks(ws, blocks)

    apply_design_block_borders(ws)
    refresh_counts(ws)
    apply_fills(ws)

    wb.save(INVENTORY_FILE)

    print(f"Sorted {len(blocks)} design block(s) alphabetically:")
    for b in blocks:
        n_variants = len(b["variants"])
        n_units    = sum(len(v["rows"]) for v in b["variants"])
        print(f"  {b['design']:<30} {n_variants} variant(s), {n_units} unit(s)")
    print(f"\nSaved → {INVENTORY_FILE}  (last data row: {last_row})")


if __name__ == "__main__":
    main()
