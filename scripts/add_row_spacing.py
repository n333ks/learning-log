#!/usr/bin/env python3
"""
Add structured blank-row spacing to inventory_master.xlsx.

  • 1 blank row between distinct variants (full 5-key match) within the same
    Design Name
  • 2 blank rows between Design Name groups
  • 1 blank row immediately after the header (row 2), so real data starts at
    row 3

Insertions are applied bottom-to-top so earlier row numbers stay valid.
Conditional-formatting ranges and design-block borders are then recalculated
to match the new layout.
"""

from openpyxl import load_workbook
from fix_inventory_formatting import fix_formatting

INVENTORY_FILE = "inventory_master.xlsx"
HEADER_ROW     = 1
OLD_DATA_START = 2   # current first data row
NEW_DATA_START = 3   # after the top-blank insertion


def scan_row_structure(ws):
    """Return [(row_num, design, variant_tuple)] for every non-blank data row."""
    rows = []
    for r in range(OLD_DATA_START, ws.max_row + 1):
        design = ws.cell(row=r, column=1).value
        if design:
            variant = tuple(
                str(ws.cell(row=r, column=c).value or "").strip()
                for c in range(1, 6)   # cols A–E
            )
            rows.append((r, design, variant))
    return rows


def compute_insertions(row_structure):
    """
    Return [(insert_before_row, n_blanks)] for every boundary that needs
    spacing.  Rules:
      same Design Name, different variant  →  1 blank
      different Design Name                →  2 blanks
    """
    insertions = []
    prev_design  = None
    prev_variant = None

    for row, design, variant in row_structure:
        if prev_design is None:
            prev_design  = design
            prev_variant = variant
            continue

        if design != prev_design:
            insertions.append((row, 2))
        elif variant != prev_variant:
            insertions.append((row, 1))

        prev_design  = design
        prev_variant = variant

    return insertions


def main():
    wb = load_workbook(INVENTORY_FILE)
    ws = wb.active

    row_structure = scan_row_structure(ws)
    insertions    = compute_insertions(row_structure)

    # Apply spacing insertions bottom-to-top (preserves earlier row numbers)
    for insert_row, n_blanks in sorted(insertions, reverse=True):
        ws.insert_rows(insert_row, amount=n_blanks)

    # Insert the top blank row at row 2 (pushing current row 2 → row 3)
    ws.insert_rows(OLD_DATA_START, amount=1)

    # Recalculate CF ranges and design-block borders for the new layout
    blocks = fix_formatting(ws, data_start=NEW_DATA_START)

    wb.save(INVENTORY_FILE)

    # ── Print results ──────────────────────────────────────────────────────────
    total_blanks = sum(n for _, n in insertions) + 1  # +1 for top blank
    variant_gaps = sum(1 for _, n in insertions if n == 1)
    design_gaps  = sum(1 for _, n in insertions if n == 2)

    print(f"Inserted {total_blanks} blank rows total:")
    print(f"  {variant_gaps} variant gap(s) × 1 blank = {variant_gaps} rows")
    print(f"  {design_gaps} design gap(s)  × 2 blanks = {design_gaps * 2} rows")
    print(f"  1 top blank (row 2 after header)")
    print(f"\nReal data now starts at row {NEW_DATA_START}.")
    print(f"Total rows in sheet: {ws.max_row}\n")

    print(f"{'Design Name':<16}  Block starts at row")
    print("─" * 38)
    for first_r, last_r, name in blocks:
        print(f"  {name:<14}  row {first_r}")

    print(f"\nSaved → {INVENTORY_FILE}")


if __name__ == "__main__":
    main()
