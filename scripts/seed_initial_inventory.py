#!/usr/bin/env python3
"""
Seed placeholder in-stock serial numbers into inventory_master.xlsx.

For every product summary row where In-Stock QTY > 0 and no serial rows
already exist beneath it, generates placeholder serials in YY-MMDD-####
format, inserts them directly below the summary row, and marks each "In Stock."

Sequence numbers climb continuously within each calendar year across all
products, never resetting between date batches.  Safe to re-run — products
that already have serial rows underneath are left untouched.
"""

from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Config (must stay in sync with create_inventory_master.py) ────────────────
INVENTORY_FILE = "inventory_master.xlsx"

GROUP_ROW  = 1   # container-group label row
HEADER_ROW = 2   # column-header row
DATA_START = 3   # first product summary row

IN_STOCK_COL    = "In-Stock QTY"
SERIAL_COL      = "Serial Number"
STATUS_COL      = "Status"
VARIANT_KEYS    = ["Design Name", "Size", "Finish", "Swing", "Glass Type"]
IN_STOCK_STATUS = "In Stock"

# Plausible past order-batch dates consumed in sheet order (one per product
# with in_stock > 0).  Pairs share a date to simulate a multi-variant batch.
# Dates within each year ascend so sequence numbers also ascend chronologically.
DATE_POOL = [
    "24-0315", "24-0315",   # Valencia 32×80, Valencia 36×80
    "24-0721",              # Valencia 36×96
    "24-1102", "24-1102",   # Sevilla 32×80, Sevilla 36×80
    "25-0210",              # Cordova 32×80
    "25-0518", "25-0518",   # Cordova 36×80, Cordova 36×96
    "25-0927", "25-0927",   # Granada 36×80, Granada 36×96
    "25-1205",              # Malaga 32×80
    "26-0103", "26-0103",   # Altea 32×80, Altea 36×80
    "26-0415",              # Altea 36×96
]


# ── Style helpers ─────────────────────────────────────────────────────────────
def _border():
    s = Side(border_style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

BORDER = _border()


# ── Column management ─────────────────────────────────────────────────────────
def build_header_map(ws):
    """Return {header_name: col_number} from HEADER_ROW; first occurrence for dupes."""
    h2c = {}
    for cell in ws[HEADER_ROW]:
        if cell.value and cell.value not in h2c:
            h2c[cell.value] = cell.column
    return h2c


def ensure_extra_columns(ws, h2c):
    """Append Serial Number and Status columns after the last column if absent."""
    # Scan the actual header row for the true last column — h2c skips duplicate
    # headers (e.g. "Presale QTY" × 3) so max(h2c.values()) would be too low.
    last_col = max(
        (cell.column for cell in ws[HEADER_ROW] if cell.value),
        default=max(h2c.values()),
    )
    next_col = last_col + 1

    for col_name in [SERIAL_COL, STATUS_COL]:
        if col_name in h2c:
            continue

        # Match the group-row fill used for non-container left columns
        gc = ws.cell(row=GROUP_ROW, column=next_col)
        gc.fill   = PatternFill("solid", fgColor="D9E2F3")
        gc.border = BORDER

        hc = ws.cell(row=HEADER_ROW, column=next_col, value=col_name)
        hc.font      = Font(name="Arial", size=10, bold=True)
        hc.fill      = PatternFill("solid", fgColor="BDD7EE")
        hc.alignment = Alignment(horizontal="center", vertical="center")
        hc.border    = BORDER

        ws.column_dimensions[get_column_letter(next_col)].width = 18
        h2c[col_name] = next_col
        print(f"  Added column '{col_name}' → {get_column_letter(next_col)}")
        next_col += 1


# ── Sheet scanning ────────────────────────────────────────────────────────────
def find_summary_rows(ws, h2c):
    """Return [(row_num, design, size, in_stock_qty)] for every product summary row."""
    design_col  = h2c["Design Name"]
    instock_col = h2c[IN_STOCK_COL]
    results = []
    for r in range(DATA_START, ws.max_row + 1):
        name = ws.cell(row=r, column=design_col).value
        if name:
            qty  = int(ws.cell(row=r, column=instock_col).value or 0)
            size = str(ws.cell(row=r, column=h2c["Size"]).value or "")
            results.append((r, str(name), size, qty))
    return results


def has_serial_rows(ws, summary_row, next_summary_row, serial_col):
    """True if any row between the two summary rows has a value in serial_col."""
    for r in range(summary_row + 1, next_summary_row):
        if ws.cell(row=r, column=serial_col).value:
            return True
    return False


# ── Serial generation ─────────────────────────────────────────────────────────
def generate_serials(assignments):
    """
    assignments: [(row, design, size, batch_date, qty), ...]

    Sorts by batch_date so sequence numbers ascend chronologically within each
    year, then returns {row: (design, size, batch_date, [serial, ...])}
    """
    year_seq = defaultdict(lambda: 1)
    result   = {}

    for row, design, size, batch_date, qty in sorted(assignments, key=lambda x: x[3]):
        year    = batch_date[:2]
        serials = [f"{batch_date}-{year_seq[year] + j:04d}" for j in range(qty)]
        year_seq[year] += qty
        result[row] = (design, size, batch_date, serials)

    return result


# ── Write serial rows ─────────────────────────────────────────────────────────
def write_serial_rows(ws, summary_row, serials, serial_col, status_col):
    """Insert len(serials) rows directly below summary_row and fill them."""
    ws.insert_rows(summary_row + 1, amount=len(serials))

    serial_font = Font(name="Arial", size=10, color="595959", italic=True)
    status_font = Font(name="Arial", size=10, color="375623")
    status_fill = PatternFill("solid", fgColor="E2EFDA")  # light green

    for offset, serial in enumerate(serials):
        r = summary_row + 1 + offset

        sc = ws.cell(row=r, column=serial_col, value=serial)
        sc.font      = serial_font
        sc.alignment = Alignment(horizontal="left", vertical="center")
        sc.border    = BORDER

        stc = ws.cell(row=r, column=status_col, value=IN_STOCK_STATUS)
        stc.font      = status_font
        stc.fill      = status_fill
        stc.alignment = Alignment(horizontal="center", vertical="center")
        stc.border    = BORDER


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    wb = load_workbook(INVENTORY_FILE)
    ws = wb.active

    h2c = build_header_map(ws)
    ensure_extra_columns(ws, h2c)

    serial_col = h2c[SERIAL_COL]
    status_col = h2c[STATUS_COL]

    summary_rows = find_summary_rows(ws, h2c)
    date_iter    = iter(DATE_POOL)
    assignments  = []
    skipped      = []

    for i, (row, design, size, qty) in enumerate(summary_rows):
        next_row = summary_rows[i + 1][0] if i + 1 < len(summary_rows) else ws.max_row + 100

        if qty <= 0:
            continue

        if has_serial_rows(ws, row, next_row, serial_col):
            skipped.append(f"{design} {size}")
            continue

        batch_date = next(date_iter)
        assignments.append((row, design, size, batch_date, qty))

    if not assignments:
        print("Nothing to seed — all in-stock products already have serial rows.")
        return

    serial_map = generate_serials(assignments)

    # Insert rows bottom-to-top so earlier row numbers aren't shifted mid-loop
    for row in sorted(serial_map.keys(), reverse=True):
        _, _, _, serials = serial_map[row]
        write_serial_rows(ws, row, serials, serial_col, status_col)

    wb.save(INVENTORY_FILE)

    # ── Summary ───────────────────────────────────────────────────────────────
    COL = 64
    print(f"\n{'─' * COL}")
    print("  SEEDING SUMMARY")
    print(f"{'─' * COL}")
    print(f"  {'Product':<28} {'Batch':^10} {'QTY':>4}   Serial range")
    print(f"  {'─'*28} {'─'*10} {'─'*4}   {'─'*28}")

    year_ranges = defaultdict(lambda: [None, None])
    total = 0

    for row in sorted(serial_map.keys()):
        design, size, batch_date, serials = serial_map[row]
        first_seq = int(serials[0].split("-")[2])
        last_seq  = int(serials[-1].split("-")[2])
        year      = batch_date[:2]

        yr    = year_ranges[year]
        yr[0] = first_seq if yr[0] is None else min(yr[0], first_seq)
        yr[1] = last_seq  if yr[1] is None else max(yr[1], last_seq)

        label = f"{design} {size}"
        print(f"  {label:<28} {batch_date:^10} {len(serials):>4}   "
              f"{serials[0]}  …  {serials[-1]}")
        total += len(serials)

    print(f"\n  Year sequence ranges used:")
    for year in sorted(year_ranges.keys()):
        lo, hi = year_ranges[year]
        print(f"    20{year}: {lo:04d} – {hi:04d}  ({hi - lo + 1} units)")

    if skipped:
        print(f"\n  Skipped (serial rows already present): {', '.join(skipped)}")

    print(f"\n  Total serials seeded: {total}  ({len(serial_map)} products)")
    print(f"  Saved → {INVENTORY_FILE}")
    print(f"{'─' * COL}\n")


if __name__ == "__main__":
    main()
