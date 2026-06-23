#!/usr/bin/env python3
"""
Regenerate inventory_master.xlsx from inventory.db.

Called at the end of every script that modifies data.
Can also be run standalone:
    python3 scripts/export_excel.py
"""

import os
import sys
import datetime

_SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _SCRIPTS_DIR)
sys.path.insert(0, os.path.join(_SCRIPTS_DIR, "legacy"))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from constants import (
    INVENTORY_FILE, SALES_TAB, WAREHOUSE_TAB,
    C_SKU, C_INSTOCK, C_INPROD, C_PRESALE, C_OPTIMAL, C_VARIANCE, C_SERIAL, C_STATUS,
    S_ORDER, S_CUSTOMER, S_DESIGN, S_SIZE, S_FINISH, S_SWING, S_GLASS,
    S_SKU, S_SERIAL, S_CONTAINER, S_DATE, S_STATUS,
    W_ORDER, W_CUSTOMER, W_DESIGN, W_SIZE, W_FINISH, W_SWING, W_GLASS,
    W_SKU, W_SERIAL, W_CONTAINER, W_DATE, W_STATUS,
    get_sku,
)
from db import get_conn, get_units_summary, get_all_units
from apply_status_fills import apply_fills, fit_columns
from fix_inventory_formatting import apply_design_block_borders

WHITE = "FFFFFFFF"
BLACK = "FF000000"
RED   = "FFFF0000"

_FILLS = {
    "In Stock":      PatternFill("solid", fgColor="C6EFCE"),
    "Pre-Sale":      PatternFill("solid", fgColor="FFEB9C"),
    "In Production": PatternFill("solid", fgColor="FFC7CE"),
    "Allocated":     PatternFill("solid", fgColor="BDD7EE"),
}
_NO_FILL = PatternFill(fill_type=None)


def _font(bold=False, color=BLACK):
    return Font(name="Arial", size=10, bold=bold, color=color)


def _center():
    return Alignment(horizontal="center")


def _left():
    return Alignment(horizontal="left")


def _status_fill(value):
    if not value:
        return _NO_FILL
    s = str(value)
    if s.startswith("Allocated"):
        return _FILLS["Allocated"]
    if s.startswith("Pre-Sale"):
        return _FILLS["Pre-Sale"]
    return _FILLS.get(s, _NO_FILL)


# ── Inventory tab ──────────────────────────────────────────────────────────────

def _write_inventory_tab(ws, conn):
    # Header row
    headers = [
        "Design Name", "Size", "Finish", "Swing", "Glass Type",
        "SKU", "In-Stock QTY", "In-Production QTY", "Pre-Sale QTY",
        "Optimal Count", "Variance", "Serial Number", "Status",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c)
        cell.value = h
        cell.font  = _font(bold=True)
        cell.alignment = _center()

    # Gather all units grouped by variant (order: design, size, finish, swing, glass)
    all_units = get_all_units(conn)
    summary   = {r["id"]: r for r in get_units_summary(conn)}

    # Group units by variant_id, preserving DB order
    from collections import OrderedDict
    variant_groups = OrderedDict()
    for u in all_units:
        vid = u["variant_id"]
        if vid not in variant_groups:
            variant_groups[vid] = []
        variant_groups[vid].append(u)

    # Build ordered list of (variant_id, design_name) for blank-row logic
    variant_list = list(variant_groups.keys())

    row = 3  # data starts at row 3 (row 2 is blank)

    prev_design = None
    for idx, vid in enumerate(variant_list):
        units   = variant_groups[vid]
        summ    = summary[vid]
        design  = summ["design_name"]

        # Spacing: 2 blank rows between designs, 1 blank between variants of same design
        if idx == 0:
            row = 3  # start here (row 2 left blank per original)
        else:
            if design != prev_design:
                row += 2
            else:
                row += 1

        prev_design = design

        for unit_idx, u in enumerate(units):
            first = (unit_idx == 0)
            key   = (u["design_name"], u["size"], u["finish"], u["swing"], u["glass_type"])

            # Columns A-E
            for c, val in enumerate(key, 1):
                cell = ws.cell(row, c)
                cell.value     = val
                cell.font      = _font(bold=True) if first else _font(color=WHITE)
                cell.alignment = _left()

            # F: SKU
            ws.cell(row, C_SKU).value     = u["sku"]
            ws.cell(row, C_SKU).font      = _font()
            ws.cell(row, C_SKU).alignment = _center()

            # G-K: counts + optimal + variance (first row only)
            if first:
                in_stock  = summ["in_stock"]
                in_prod   = summ["in_prod"]
                pre_sale  = summ["pre_sale"]
                optimal   = summ["optimal_count"]
                variance  = summ["variance"]

                for col, val in [
                    (C_INSTOCK,  in_stock),
                    (C_INPROD,   in_prod),
                    (C_PRESALE,  pre_sale),
                    (C_OPTIMAL,  optimal),
                    (C_VARIANCE, variance),
                ]:
                    cell = ws.cell(row, col)
                    cell.value     = val
                    cell.alignment = _center()
                    cell.font      = _font(color="FFFF0000") if (col == C_VARIANCE and variance < 0) else _font()
            else:
                for col in (C_INSTOCK, C_INPROD, C_PRESALE, C_OPTIMAL, C_VARIANCE):
                    ws.cell(row, col).value = None

            # L: Serial
            ws.cell(row, C_SERIAL).value     = u["serial_number"]
            ws.cell(row, C_SERIAL).font      = _font()
            ws.cell(row, C_SERIAL).alignment = _center()

            # M: Status + fill
            status_val = u["status"]
            status_cell = ws.cell(row, C_STATUS)
            status_cell.value     = status_val
            status_cell.font      = _font()
            status_cell.alignment = _center()
            status_cell.fill      = _status_fill(status_val)

            row += 1

    return row - 1  # last written row


# ── Sales tab ──────────────────────────────────────────────────────────────────

def _write_sales_tab(ws, conn):
    headers = [
        "Order #", "Customer", "Design Name", "Size", "Finish", "Swing",
        "Glass Type", "SKU", "Serial #", "Container", "Date Allocated", "Status",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c)
        cell.value = h
        cell.font  = _font(bold=True)
        cell.alignment = _center()

    rows = conn.execute("""
        SELECT so.order_number, so.customer, v.design_name, v.size, v.finish, v.swing,
               v.glass_type, v.sku, so.serial_number, so.container_id,
               so.date_allocated, so.status
        FROM sales_orders so
        JOIN variants v ON v.id = so.variant_id
        ORDER BY so.id
    """).fetchall()

    for r_idx, row in enumerate(rows, 2):
        text_cols = {
            S_ORDER:    row["order_number"],
            S_CUSTOMER: row["customer"],
            S_DESIGN:   row["design_name"],
            S_SIZE:     row["size"],
            S_FINISH:   row["finish"],
            S_SWING:    row["swing"],
            S_GLASS:    row["glass_type"],
        }
        center_cols = {
            S_SKU:       row["sku"],
            S_SERIAL:    row["serial_number"],
            S_CONTAINER: row["container_id"],
            S_DATE:      row["date_allocated"],
            S_STATUS:    row["status"],
        }
        for col, val in text_cols.items():
            cell = ws.cell(r_idx, col)
            cell.value     = val
            cell.font      = _font()
            cell.alignment = _left()
        for col, val in center_cols.items():
            cell = ws.cell(r_idx, col)
            cell.value     = val
            cell.font      = _font()
            cell.alignment = _center()


# ── Warehouse tab ──────────────────────────────────────────────────────────────

def _write_warehouse_tab(ws, conn):
    headers = [
        "Order #", "Customer", "Design Name", "Size", "Finish", "Swing",
        "Glass Type", "SKU", "Serial #", "Container", "Date Arrived", "Status",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c)
        cell.value = h
        cell.font  = _font(bold=True)
        cell.alignment = _center()

    rows = conn.execute("""
        SELECT wh.order_number, wh.customer, v.design_name, v.size, v.finish, v.swing,
               v.glass_type, v.sku, wh.serial_number, wh.container_id,
               wh.date_arrived, wh.status
        FROM warehouse wh
        JOIN variants v ON v.id = wh.variant_id
        ORDER BY wh.id
    """).fetchall()

    for r_idx, row in enumerate(rows, 2):
        text_cols = {
            W_ORDER:    row["order_number"],
            W_CUSTOMER: row["customer"],
            W_DESIGN:   row["design_name"],
            W_SIZE:     row["size"],
            W_FINISH:   row["finish"],
            W_SWING:    row["swing"],
            W_GLASS:    row["glass_type"],
        }
        center_cols = {
            W_SKU:       row["sku"],
            W_SERIAL:    row["serial_number"],
            W_CONTAINER: row["container_id"],
            W_DATE:      row["date_arrived"],
            W_STATUS:    row["status"],
        }
        for col, val in text_cols.items():
            cell = ws.cell(r_idx, col)
            cell.value     = val
            cell.font      = _font()
            cell.alignment = _left()
        for col, val in center_cols.items():
            cell = ws.cell(r_idx, col)
            cell.value     = val
            cell.font      = _font()
            cell.alignment = _center()


# ── Main export function ───────────────────────────────────────────────────────

def export_excel(conn=None):
    """Regenerate inventory_master.xlsx from inventory.db."""
    owns_conn = conn is None
    if owns_conn:
        conn = get_conn()

    wb = Workbook()

    # Rename default sheet to Inventory Master
    ws_inv = wb.active
    ws_inv.title = "Inventory Master"

    ws_sales = wb.create_sheet(SALES_TAB)
    ws_wh    = wb.create_sheet(WAREHOUSE_TAB)

    print("  Writing Inventory tab ...")
    last_row = _write_inventory_tab(ws_inv, conn)
    apply_design_block_borders(ws_inv)
    apply_fills(ws_inv)

    print("  Writing Sales tab ...")
    _write_sales_tab(ws_sales, conn)
    fit_columns(ws_sales)

    print("  Writing Warehouse tab ...")
    _write_warehouse_tab(ws_wh, conn)
    fit_columns(ws_wh)

    wb.save(INVENTORY_FILE)
    print(f"  Saved → {INVENTORY_FILE}  (last inventory row: {last_row})")

    if owns_conn:
        conn.close()


def main():
    print("Exporting inventory.db → inventory_master.xlsx ...")
    export_excel()
    print("Done.")


if __name__ == "__main__":
    main()
