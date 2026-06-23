#!/usr/bin/env python3
"""
One-time migration from inventory_master.xlsx → inventory.db.

Reads:
  - Inventory tab (row 3+): variant + unit rows
  - Sales tab (row 2+):     sales_orders rows
  - Warehouse tab (row 2+): warehouse rows

Idempotent: safe to run twice (INSERT OR IGNORE on variants;
duplicate unit detection skips already-present serials).

Usage:
    python3 scripts/migrate.py
"""

import os
import sys

# Allow running from any directory
_SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _SCRIPTS_DIR)

from openpyxl import load_workbook

from constants import (
    INVENTORY_FILE, get_sku,
    DATA_START, C_SERIAL, C_STATUS, C_OPTIMAL,
    SALES_TAB, WAREHOUSE_TAB,
    S_ORDER, S_CUSTOMER, S_DESIGN, S_SIZE, S_FINISH, S_SWING, S_GLASS,
    S_SERIAL, S_CONTAINER, S_DATE,
    W_ORDER, W_CUSTOMER, W_DESIGN, W_SIZE, W_FINISH, W_SWING, W_GLASS,
    W_SERIAL, W_CONTAINER, W_DATE, W_STATUS,
)
from db import DB_PATH, init_db, get_conn, get_or_create_variant, set_optimal_count


def _str(v):
    return str(v).strip() if v is not None else ""


def migrate_inventory(conn, ws):
    """Migrate the Inventory tab. Returns (n_variants_new, n_units)."""
    # Track first-seen row per variant so we can read optimal_count from it
    seen_variants = {}   # variant_key -> variant_id
    n_variants = 0
    n_units = 0

    # Collect existing serials to avoid duplicates
    existing_serials = set(
        r[0] for r in conn.execute("SELECT serial_number FROM units WHERE serial_number IS NOT NULL")
    )

    for r in range(DATA_START, ws.max_row + 1):
        key = tuple(_str(ws.cell(r, c).value) for c in range(1, 6))
        if not any(key):
            continue

        design, size, finish, swing, glass = key
        serial = _str(ws.cell(r, C_SERIAL).value)
        status = _str(ws.cell(r, C_STATUS).value)

        if not serial and not status:
            continue  # skip ghost rows

        sku = get_sku(key)

        if key not in seen_variants:
            # Read optimal_count from this first row of the variant
            optimal_raw = ws.cell(r, C_OPTIMAL).value
            optimal = int(optimal_raw) if optimal_raw is not None else 0

            vid = get_or_create_variant(conn, design, size, finish, swing, glass, sku)
            seen_variants[key] = vid
            set_optimal_count(conn, vid, optimal)
            n_variants += 1
        else:
            vid = seen_variants[key]

        # Skip if serial already in DB (idempotency)
        if serial and serial in existing_serials:
            continue

        # Insert unit
        conn.execute(
            "INSERT INTO units (variant_id, serial_number, status) VALUES (?,?,?)",
            (vid, serial if serial else None, status if status else "In Stock"),
        )
        if serial:
            existing_serials.add(serial)
        n_units += 1

    conn.commit()
    return n_variants, n_units


def migrate_sales(conn, ws_sales):
    """Migrate the Sales tab. Returns n_sales."""
    n_sales = 0

    # Existing sales to avoid duplicates (order_number + serial)
    existing = set(
        (r[0], r[1]) for r in conn.execute("SELECT order_number, serial_number FROM sales_orders")
    )

    for r in range(2, ws_sales.max_row + 1):
        order    = _str(ws_sales.cell(r, S_ORDER).value)
        customer = _str(ws_sales.cell(r, S_CUSTOMER).value)
        design   = _str(ws_sales.cell(r, S_DESIGN).value)
        size     = _str(ws_sales.cell(r, S_SIZE).value)
        finish   = _str(ws_sales.cell(r, S_FINISH).value)
        swing    = _str(ws_sales.cell(r, S_SWING).value)
        glass    = _str(ws_sales.cell(r, S_GLASS).value)
        serial   = _str(ws_sales.cell(r, S_SERIAL).value)
        container= _str(ws_sales.cell(r, S_CONTAINER).value)
        date_val = ws_sales.cell(r, S_DATE).value
        date_str = _str(date_val) if date_val else None

        if not order and not serial:
            continue
        if (order, serial) in existing:
            continue

        key = (design, size, finish, swing, glass)
        sku = get_sku(key)
        vid = get_or_create_variant(conn, design, size, finish, swing, glass, sku)

        conn.execute(
            "INSERT INTO sales_orders (order_number, customer, variant_id, serial_number, container_id, date_allocated, status) VALUES (?,?,?,?,?,?,'Allocated')",
            (order, customer, vid, serial, container, date_str),
        )
        existing.add((order, serial))
        n_sales += 1

    conn.commit()
    return n_sales


def migrate_warehouse(conn, ws_wh):
    """Migrate the Warehouse tab. Returns n_warehouse."""
    n_wh = 0

    existing = set(
        (r[0], r[1]) for r in conn.execute("SELECT order_number, serial_number FROM warehouse")
    )

    for r in range(2, ws_wh.max_row + 1):
        order    = _str(ws_wh.cell(r, W_ORDER).value)
        customer = _str(ws_wh.cell(r, W_CUSTOMER).value)
        design   = _str(ws_wh.cell(r, W_DESIGN).value)
        size     = _str(ws_wh.cell(r, W_SIZE).value)
        finish   = _str(ws_wh.cell(r, W_FINISH).value)
        swing    = _str(ws_wh.cell(r, W_SWING).value)
        glass    = _str(ws_wh.cell(r, W_GLASS).value)
        serial   = _str(ws_wh.cell(r, W_SERIAL).value)
        container= _str(ws_wh.cell(r, W_CONTAINER).value)
        date_val = ws_wh.cell(r, W_DATE).value
        date_str = _str(date_val) if date_val else None
        status   = _str(ws_wh.cell(r, W_STATUS).value) or "In Prep"

        if not order and not serial:
            continue
        if (order, serial) in existing:
            continue

        key = (design, size, finish, swing, glass)
        sku = get_sku(key)
        vid = get_or_create_variant(conn, design, size, finish, swing, glass, sku)

        conn.execute(
            "INSERT INTO warehouse (order_number, customer, variant_id, serial_number, container_id, date_arrived, status) VALUES (?,?,?,?,?,?,?)",
            (order, customer, vid, serial, container, date_str, status),
        )
        existing.add((order, serial))
        n_wh += 1

    conn.commit()
    return n_wh


def main():
    print(f"Initialising database at {DB_PATH} ...")
    init_db()

    print(f"Opening workbook: {INVENTORY_FILE}")
    wb = load_workbook(INVENTORY_FILE)
    ws       = wb.active
    ws_sales = wb[SALES_TAB]
    ws_wh    = wb[WAREHOUSE_TAB]

    conn = get_conn()

    print("\nMigrating Inventory tab ...")
    n_variants, n_units = migrate_inventory(conn, ws)
    print(f"  {n_variants} variant(s), {n_units} unit(s) migrated.")

    print("\nMigrating Sales tab ...")
    n_sales = migrate_sales(conn, ws_sales)
    print(f"  {n_sales} sales order row(s) migrated.")

    print("\nMigrating Warehouse tab ...")
    n_wh = migrate_warehouse(conn, ws_wh)
    print(f"  {n_wh} warehouse row(s) migrated.")

    conn.close()

    print(f"\nMigration complete.")
    print(f"  Variants:       {n_variants}")
    print(f"  Units:          {n_units}")
    print(f"  Sales orders:   {n_sales}")
    print(f"  Warehouse rows: {n_wh}")


if __name__ == "__main__":
    main()
