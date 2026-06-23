#!/usr/bin/env python3
"""
Rebuild inventory_master.xlsx with a flat per-unit row structure.

Steps
-----
1. Reads the current file.  Extracts every In Stock serial and the product
   variant (Design Name / Size / Finish / Swing / Glass Type) it belongs to,
   plus each variant's Optimal Count.
2. Writes a new inventory_master.xlsx where EVERY row is one physical unit:
     • Variant fields repeat on every row
     • In-Stock / In-Production / Pre-Sale QTY are live COUNTIFS formulas
     • Variance = In-Stock + In-Production + Pre-Sale − Optimal Count
     • Serial Number and Status are per unit
3. Row order within each variant: In Stock → Pre-Sale → In Production
   (Pre-Sale and In Production rows: none yet — added as units arrive)
4. Variance column CF: red fill + red font when negative.
5. Status column (L) CF: coloured fill, black font — col L only, no row bleed.
6. Design-block borders: medium outer border around each Design Name's rows.
"""

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from fix_inventory_formatting import apply_design_block_borders
from apply_status_fills import apply_fills

INVENTORY_FILE = "inventory_master.xlsx"

# ── Old-file read config ───────────────────────────────────────────────────────
OLD_HEADER_ROW   = 2
OLD_DATA_START   = 3
OLD_VARIANT_KEYS = ["Design Name", "Size", "Finish", "Swing", "Glass Type"]
OLD_OPTIMAL_COL  = "Optimal Count"
OLD_SERIAL_COL   = "Serial Number"
OLD_STATUS_COL   = "Status"
IN_STOCK_STATUS  = "In Stock"

# ── New-file column layout ─────────────────────────────────────────────────────
C_DESIGN   = 1   # A  Design Name
C_SIZE     = 2   # B  Size
C_FINISH   = 3   # C  Finish
C_SWING    = 4   # D  Swing
C_GLASS    = 5   # E  Glass Type
C_INSTOCK  = 6   # F  In-Stock QTY       (COUNTIFS)
C_INPROD   = 7   # G  In-Production QTY  (COUNTIFS)
C_PRESALE  = 8   # H  Pre-Sale QTY       (COUNTIFS)
C_OPTIMAL  = 9   # I  Optimal Count
C_VARIANCE = 10  # J  Variance
C_SERIAL   = 11  # K  Serial Number
C_STATUS   = 12  # L  Status
TOTAL_COLS = C_STATUS

HEADER_ROW = 1
DATA_START = 2

HEADERS = [
    "Design Name", "Size", "Finish", "Swing", "Glass Type",
    "In-Stock QTY", "In-Production QTY", "Pre-Sale QTY",
    "Optimal Count", "Variance", "Serial Number", "Status",
]

COL_WIDTHS = {
    C_DESIGN:   16,
    C_SIZE:     12,
    C_FINISH:   21,
    C_SWING:    16,
    C_GLASS:    13,
    C_INSTOCK:  14,
    C_INPROD:   20,
    C_PRESALE:  13,
    C_OPTIMAL:  14,
    C_VARIANCE: 11,
    C_SERIAL:   18,
    C_STATUS:   16,
}

# Status CF: (excel_comparison_value, CellIsRule operator, fill ARGB).
# Applied to col L only; font always black.  8-char ARGB = fully opaque.
STATUS_CF = [
    ('"In Stock"',      "equal", "FFC6EFCE"),   # green
    ('"Pre-Sale"',      "equal", "FFFFEB9C"),   # yellow
    ('"In Production"', "equal", "FFFFC7CE"),   # red (light)
]
ALLOCATED_FILL    = "FFBDD7EE"   # blue — prefix-match FormulaRule
VARIANCE_NEG_FILL = "FFFFCCCC"
VARIANCE_NEG_FONT = "FFC00000"

# Columns that get center alignment in data rows
CENTER_COLS = {C_INSTOCK, C_INPROD, C_PRESALE, C_OPTIMAL, C_VARIANCE, C_SERIAL, C_STATUS}


# ── Extract data from current file ────────────────────────────────────────────
def extract_old_data(filepath):
    """
    Returns:
        variants_ordered : list of variant tuples in original sheet order
        variant_optimal  : {variant_tuple: int}
        variant_serials  : {variant_tuple: [serial_str, ...]}   (In Stock only)
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    h2c = {}
    for cell in ws[OLD_HEADER_ROW]:
        if cell.value and cell.value not in h2c:
            h2c[cell.value] = cell.column

    for col_name in OLD_VARIANT_KEYS + [OLD_OPTIMAL_COL]:
        if col_name not in h2c:
            raise ValueError(f"Current file missing expected column: {col_name!r}")

    variant_cols    = [h2c[k] for k in OLD_VARIANT_KEYS]
    optimal_col_idx = h2c[OLD_OPTIMAL_COL]
    serial_col_idx  = h2c.get(OLD_SERIAL_COL)
    status_col_idx  = h2c.get(OLD_STATUS_COL)

    variants_ordered = []
    variant_optimal  = {}
    variant_serials  = {}
    current_variant  = None

    for r in range(OLD_DATA_START, ws.max_row + 1):
        design = ws.cell(row=r, column=variant_cols[0]).value

        if design:
            vkey = tuple(
                str(ws.cell(row=r, column=c).value or "").strip()
                for c in variant_cols
            )
            optimal = int(ws.cell(row=r, column=optimal_col_idx).value or 0)
            if vkey not in variant_optimal:
                variants_ordered.append(vkey)
                variant_optimal[vkey] = optimal
                variant_serials[vkey] = []
            current_variant = vkey

        elif current_variant and serial_col_idx:
            serial = ws.cell(row=r, column=serial_col_idx).value
            status = ws.cell(row=r, column=status_col_idx).value if status_col_idx else None
            if serial and status == IN_STOCK_STATUS:
                variant_serials[current_variant].append(str(serial).strip())

    return variants_ordered, variant_optimal, variant_serials


# ── Style helpers ─────────────────────────────────────────────────────────────
def _border():
    s = Side(border_style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

BORDER    = _border()
FONT_HDR  = Font(name="Arial", size=10, bold=True,  color="FF000000")
FONT_BASE = Font(name="Arial", size=10, bold=False, color="FF000000")
FILL_HDR  = PatternFill("solid", fgColor="BDD7EE")
ALIGN_C   = Alignment(horizontal="center", vertical="center")
ALIGN_L   = Alignment(horizontal="left",   vertical="center")


# ── Build new workbook ────────────────────────────────────────────────────────
def build_new_workbook(variants_ordered, variant_optimal, variant_serials):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Master"

    # Header row
    for col, header in enumerate(HEADERS, start=1):
        cell           = ws.cell(row=HEADER_ROW, column=col, value=header)
        cell.font      = FONT_HDR
        cell.fill      = FILL_HDR
        cell.alignment = ALIGN_C
        cell.border    = BORDER
    ws.row_dimensions[HEADER_ROW].height = 20

    current_row = DATA_START

    for vkey in variants_ordered:
        design, size, finish, swing, glass = vkey
        optimal = variant_optimal[vkey]
        serials = variant_serials[vkey]

        # Row order: In Stock first, then Pre-Sale, then In Production.
        # Pre-Sale and In Production rows are blank for now — added as units arrive.
        unit_rows = [(s, IN_STOCK_STATUS) for s in serials]

        for serial, status in unit_rows:
            r = current_row

            row_values = [
                design, size, finish, swing, glass,
                f'=COUNTIFS($A:$A,A{r},$B:$B,B{r},$C:$C,C{r},$D:$D,D{r},$E:$E,E{r},$L:$L,"In Stock")',
                f'=COUNTIFS($A:$A,A{r},$B:$B,B{r},$C:$C,C{r},$D:$D,D{r},$E:$E,E{r},$L:$L,"In Production")',
                f'=COUNTIFS($A:$A,A{r},$B:$B,B{r},$C:$C,C{r},$D:$D,D{r},$E:$E,E{r},$L:$L,"Pre-Sale")',
                optimal,
                f"=F{r}+G{r}+H{r}-I{r}",
                serial,
                status,
            ]

            for col, value in enumerate(row_values, start=1):
                cell           = ws.cell(row=r, column=col, value=value)
                cell.font      = FONT_BASE
                cell.border    = BORDER
                cell.alignment = ALIGN_C if col in CENTER_COLS else ALIGN_L

            ws.row_dimensions[r].height = 16
            current_row += 1

    last_data_row = current_row - 1

    if last_data_row >= DATA_START:
        var_range    = f"J{DATA_START}:J{last_data_row}"
        status_range = f"L{DATA_START}:L{last_data_row}"

        # Variance CF: negative → red fill + red font
        ws.conditional_formatting.add(
            var_range,
            CellIsRule(
                operator="lessThan",
                formula=["0"],
                fill=PatternFill("solid", fgColor=VARIANCE_NEG_FILL),
                font=Font(name="Arial", size=10, color=VARIANCE_NEG_FONT),
            ),
        )

        # Status CF: coloured fill, black font — col L only
        for excel_value, operator, fill_argb in STATUS_CF:
            ws.conditional_formatting.add(
                status_range,
                CellIsRule(
                    operator=operator,
                    formula=[excel_value],
                    fill=PatternFill("solid", fgColor=fill_argb),
                    font=Font(name="Arial", size=10, color="FF000000"),
                ),
            )
        ws.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[f'=LEFT($L{DATA_START},9)="Allocated"'],
                fill=PatternFill("solid", fgColor=ALLOCATED_FILL),
                font=Font(name="Arial", size=10, color="FF000000"),
            ),
        )

    # Column widths
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze row 1 (header) and col A (Design Name): freeze_panes = B2
    ws.freeze_panes = ws.cell(row=DATA_START, column=C_SIZE)

    # Design-block borders (medium outer border per Design Name group)
    apply_design_block_borders(ws)

    apply_fills(ws)
    wb.save(INVENTORY_FILE)
    return last_data_row


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print(f"Reading {INVENTORY_FILE} ...")
    variants, optimal, serials = extract_old_data(INVENTORY_FILE)
    total_extracted = sum(len(s) for s in serials.values())
    print(f"  {len(variants)} product variants found, "
          f"{total_extracted} In Stock serials to migrate.\n")

    print("Building new file ...")
    last_row = build_new_workbook(variants, optimal, serials)
    total_rows = last_row - DATA_START + 1

    # ── Summary ───────────────────────────────────────────────────────────────
    W = 70
    print(f"\n{'═' * W}")
    print("  MIGRATION SUMMARY")
    print(f"{'═' * W}")
    print(f"  {'Product':<32} {'Serials':>7}   Range")
    print(f"  {'─'*32} {'─'*7}   {'─'*26}")

    total_out = 0
    for vkey in variants:
        design, size = vkey[0], vkey[1]
        slist  = serials[vkey]
        count  = len(slist)
        total_out += count
        label  = f"{design} {size}"
        if count > 1:
            rng = f"{slist[0]}  …  {slist[-1]}"
        elif count == 1:
            rng = slist[0]
        else:
            rng = "—  (0 units, not written to new file)"
        print(f"  {label:<32} {count:>5}     {rng}")

    print(f"\n  {'─'*32} {'─'*7}")
    print(f"  {'TOTAL':<32} {total_out:>5}")
    print(f"\n  Data rows written : {total_rows}")
    print(f"  Saved             → {INVENTORY_FILE}")
    print(f"{'═' * W}\n")


if __name__ == "__main__":
    main()
