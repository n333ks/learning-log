#!/usr/bin/env python3
"""
Scratch test: append two Pre-Sale rows to verify yellow fill renders correctly.
Run once, open Excel to check, then delete the rows (or run git checkout to revert).
"""

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from apply_status_fills import apply_fills

INVENTORY_FILE = "inventory_master.xlsx"

TEST_ROWS = [
    # (Design Name, Size, Finish, Swing, Glass Type, Serial, Status)
    ("Valencia", '32" x 80"', "Matte Black", "Left Inswing", "Clear",  "TEST-0001", "Pre-Sale"),
    ("Altea",    '36" x 80"', "Brushed Nickel", "Right Inswing", "Frosted", "TEST-0002", "Pre-Sale"),
]

WHITE = "FFFFFFFF"

def main():
    wb = load_workbook(INVENTORY_FILE)
    ws = wb.active

    next_row = ws.max_row + 2  # leave one blank row gap

    for design, size, finish, swing, glass, serial, status in TEST_ROWS:
        r = next_row
        ws.cell(r, 1).value  = design
        ws.cell(r, 2).value  = size
        ws.cell(r, 3).value  = finish
        ws.cell(r, 4).value  = swing
        ws.cell(r, 5).value  = glass
        ws.cell(r, 11).value = serial
        ws.cell(r, 12).value = status

        # Match detail-row style: variant text in white font, K/L centered
        for c in range(1, 6):
            cell = ws.cell(r, c)
            cell.font = Font(
                name=cell.font.name or "Arial",
                size=cell.font.size or 10,
                color=WHITE,
            )
        for c in (11, 12):
            ws.cell(r, c).alignment = Alignment(horizontal="center")

        print(f"  Added row {r}: {design} | {serial} | {status}")
        next_row += 1

    apply_fills(ws)
    wb.save(INVENTORY_FILE)
    print(f"\nSaved → {INVENTORY_FILE}")
    print("Open in Excel — Pre-Sale rows should be yellow.")
    print("To clean up: run  git checkout inventory_master.xlsx")

if __name__ == "__main__":
    main()
