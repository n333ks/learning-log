#!/usr/bin/env python3
"""
For every detail row in inventory_master.xlsx (same variant as the row above,
not a blank spacer):
  1. Set font color to white on cols A–E (Design Name … Glass Type) so the
     repeated text is invisible but values remain intact for formula references.
  2. Clear the Optimal Count cell (col I) — value only appears on summary rows.

Summary rows (first row of each variant group) are left untouched.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font

INVENTORY_FILE = "inventory_master.xlsx"
DATA_START     = 3

C_GLASS   = 5    # last variant-label column
C_OPTIMAL = 9    # I
C_INSTOCK  = 6   # F — formula columns (verified only)
C_INPROD   = 7   # G
C_PRESALE  = 8   # H
C_VARIANCE = 10  # J

WHITE = "FFFFFFFF"


def main():
    wb = load_workbook(INVENTORY_FILE)
    ws = wb.active

    prev_variant    = None
    detail_count    = 0
    summary_count   = 0
    summary_rows    = []   # (row, variant) — for formula verification

    for r in range(DATA_START, ws.max_row + 1):
        design = ws.cell(row=r, column=1).value

        if not design:
            prev_variant = None
            continue

        variant = tuple(
            str(ws.cell(row=r, column=c).value or "").strip()
            for c in range(1, 6)
        )

        if variant == prev_variant:
            # ── Detail row ────────────────────────────────────────────────────
            # 1. White font on cols A–E (value stays, text becomes invisible)
            for c in range(1, C_GLASS + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = Font(
                    name  = cell.font.name  or "Arial",
                    size  = cell.font.size  or 10,
                    bold  = cell.font.bold,
                    color = WHITE,
                )
            # 2. Clear Optimal Count
            ws.cell(row=r, column=C_OPTIMAL).value = None
            detail_count += 1
        else:
            # ── Summary row ───────────────────────────────────────────────────
            summary_rows.append((r, variant))
            summary_count += 1

        prev_variant = variant

    wb.save(INVENTORY_FILE)

    # ── Report ─────────────────────────────────────────────────────────────────
    print(f"Summary rows (unchanged) : {summary_count}")
    print(f"Detail rows updated      : {detail_count}")
    print(f"  • Cols A–E font → white on all {detail_count} detail rows")
    print(f"  • Optimal Count cleared on all {detail_count} detail rows\n")

    # ── Verify formula columns untouched ───────────────────────────────────────
    wb2 = load_workbook(INVENTORY_FILE)
    ws2 = wb2.active

    formula_on_detail = []
    prev_variant = None

    for r in range(DATA_START, ws2.max_row + 1):
        design = ws2.cell(row=r, column=1).value
        if not design:
            prev_variant = None
            continue

        variant = tuple(
            str(ws2.cell(row=r, column=c).value or "").strip()
            for c in range(1, 6)
        )

        if variant == prev_variant:
            # Detail row — formula cols should be blank
            for c, label in ((C_INSTOCK, "F"), (C_INPROD, "G"),
                             (C_PRESALE, "H"), (C_VARIANCE, "J")):
                val = ws2.cell(row=r, column=c).value
                if val is not None:
                    formula_on_detail.append((r, label, val))

        prev_variant = variant

    if formula_on_detail:
        print("WARNING: formula values found on detail rows:")
        for r, col, val in formula_on_detail:
            print(f"  row {r} col {col}: {val}")
    else:
        print("Verification: F/G/H/J still blank on all detail rows ✓")

    # Spot-check one summary row still has its formula
    sr, sv = summary_rows[0]
    f_val = ws2.cell(row=sr, column=C_INSTOCK).value
    print(f"Verification: summary row {sr} F={f_val} ✓" if f_val else
          f"WARNING: summary row {sr} F is unexpectedly blank")

    print(f"\nSaved → {INVENTORY_FILE}")


if __name__ == "__main__":
    main()
