#!/usr/bin/env python3
"""
Fix inventory_master.xlsx formula placement.

Two problems this resolves:
  1. COUNTIFS and Variance formulas are on every row of a variant group.
     They should only be on the first (summary) row.
  2. After add_row_spacing.py inserted rows, formula cell references are
     off by one (e.g. the formula on row 3 says A2, not A3) because
     openpyxl does not rewrite formula strings when insert_rows is called.

Fix applied:
  - First row of each variant group: rewrite F, G, H, J formulas so they
    self-reference the correct row number.
  - All other rows in the same variant group: clear F, G, H, J entirely.

Variant groups are delimited by blank rows (col A empty).
"""

from openpyxl import load_workbook

INVENTORY_FILE = "inventory_master.xlsx"
DATA_START     = 3   # row 1 = header, row 2 = top blank

C_INSTOCK  = 6    # F
C_INPROD   = 7    # G
C_PRESALE  = 8    # H
C_VARIANCE = 10   # J


def instock_formula(r):
    return (f'=COUNTIFS($A:$A,A{r},$B:$B,B{r},$C:$C,C{r},'
            f'$D:$D,D{r},$E:$E,E{r},$L:$L,"In Stock")')


def inprod_formula(r):
    return (f'=COUNTIFS($A:$A,A{r},$B:$B,B{r},$C:$C,C{r},'
            f'$D:$D,D{r},$E:$E,E{r},$L:$L,"In Production")')


def presale_formula(r):
    return (f'=COUNTIFS($A:$A,A{r},$B:$B,B{r},$C:$C,C{r},'
            f'$D:$D,D{r},$E:$E,E{r},$L:$L,"Pre-Sale")')


def variance_formula(r):
    return f'=F{r}+G{r}+H{r}-I{r}'


def main():
    wb = load_workbook(INVENTORY_FILE)
    ws = wb.active

    prev_variant  = None
    summary_count = 0
    cleared_count = 0
    summary_rows  = []   # (row_num, variant) for verification

    for r in range(DATA_START, ws.max_row + 1):
        design = ws.cell(row=r, column=1).value

        if not design:
            prev_variant = None   # blank spacer — reset variant tracking
            continue

        variant = tuple(
            str(ws.cell(row=r, column=c).value or "").strip()
            for c in range(1, 6)   # cols A–E
        )

        if variant != prev_variant:
            # ── Summary row: rewrite formulas with the correct row reference ──
            ws.cell(row=r, column=C_INSTOCK ).value = instock_formula(r)
            ws.cell(row=r, column=C_INPROD  ).value = inprod_formula(r)
            ws.cell(row=r, column=C_PRESALE ).value = presale_formula(r)
            ws.cell(row=r, column=C_VARIANCE).value = variance_formula(r)
            summary_rows.append((r, variant))
            summary_count += 1
        else:
            # ── Detail row: clear formula columns ────────────────────────────
            for c in (C_INSTOCK, C_INPROD, C_PRESALE, C_VARIANCE):
                ws.cell(row=r, column=c).value = None
            cleared_count += 1

        prev_variant = variant

    wb.save(INVENTORY_FILE)

    # ── Report ─────────────────────────────────────────────────────────────────
    print(f"Summary rows updated : {summary_count}")
    print(f"Detail rows cleared  : {cleared_count}\n")

    # ── Verify: Valencia 32×80 (first variant group) ───────────────────────────
    sample_row, sample_variant = summary_rows[0]
    design_str = f"{sample_variant[0]} {sample_variant[1]}"
    print(f"Verification — {design_str}")
    print(f"  Summary row  : {sample_row}")

    # Re-read to confirm formulas
    wb2 = load_workbook(INVENTORY_FILE)
    ws2 = wb2.active

    formula_written = ws2.cell(row=sample_row, column=C_INSTOCK).value
    print(f"  F{sample_row} formula : {formula_written}")

    # Count actual In Stock rows for the sample variant
    actual_instock = sum(
        1
        for r in range(DATA_START, ws2.max_row + 1)
        if ws2.cell(row=r, column=1).value
        and tuple(
            str(ws2.cell(row=r, column=c).value or "").strip()
            for c in range(1, 6)
        ) == sample_variant
        and ws2.cell(row=r, column=12).value == "In Stock"
    )
    print(f"  Actual 'In Stock' rows for this variant: {actual_instock}")

    # Confirm the formula argument matches the row (self-reference check)
    expected_formula = instock_formula(sample_row)
    if formula_written == expected_formula:
        print(f"  Formula references row {sample_row} correctly ✓")
    else:
        print(f"  WARNING: formula mismatch")
        print(f"    expected : {expected_formula}")
        print(f"    actual   : {formula_written}")

    # Confirm detail rows in the group have no value in col F
    detail_non_blank = [
        r for r in range(DATA_START, ws2.max_row + 1)
        if r != sample_row
        and ws2.cell(row=r, column=1).value
        and tuple(
            str(ws2.cell(row=r, column=c).value or "").strip()
            for c in range(1, 6)
        ) == sample_variant
        and ws2.cell(row=r, column=C_INSTOCK).value is not None
    ]
    if detail_non_blank:
        print(f"  WARNING: {len(detail_non_blank)} detail row(s) still have F values: {detail_non_blank}")
    else:
        print(f"  Detail rows have blank In-Stock QTY (F) ✓")

    print(f"\nSaved → {INVENTORY_FILE}")


if __name__ == "__main__":
    main()
