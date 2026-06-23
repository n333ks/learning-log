#!/usr/bin/env python3
"""
Apply direct background fills to Status cells (col L) based on value.
Can be imported as a module (apply_fills) or run standalone.

  In Stock       → green  (#C6EFCE)
  Pre-Sale       → yellow (#FFEB9C)
  In Production  → red    (#FFC7CE)
  Allocated *    → blue   (#BDD7EE)  (prefix match)
  anything else  → no fill
"""

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from constants import C_STATUS, DATA_START, INVENTORY_FILE, SALES_TAB, WAREHOUSE_TAB

_FILLS = {
    "In Stock":      PatternFill("solid", fgColor="C6EFCE"),
    "Pre-Sale":      PatternFill("solid", fgColor="FFEB9C"),
    "In Production": PatternFill("solid", fgColor="FFC7CE"),
    "Allocated":     PatternFill("solid", fgColor="BDD7EE"),
}
_NO_FILL = PatternFill(fill_type=None)


def _status_fill(value):
    if not value:
        return _NO_FILL
    s = str(value)
    if s.startswith("Allocated"):
        return _FILLS["Allocated"]
    if s.startswith("Pre-Sale"):
        return _FILLS["Pre-Sale"]
    return _FILLS.get(s, _NO_FILL)


def fit_columns(ws):
    """Auto-fit every column width to its longest cell content."""
    from openpyxl.utils import get_column_letter
    col_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                col_widths[cell.column] = max(col_widths.get(cell.column, 0), len(str(cell.value)))
    for col, max_len in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = max(max_len + 4, 10)


def apply_fills(ws):
    """Apply status fills to every cell in column M of the given worksheet."""
    fit_columns(ws)
    for r in range(DATA_START, ws.max_row + 1):
        cell = ws.cell(row=r, column=C_STATUS)
        cell.fill = _status_fill(cell.value)
        cell.alignment = Alignment(horizontal="center")
        if cell.value:
            cell.font = Font(
                name=cell.font.name or "Arial",
                size=cell.font.size or 10,
                bold=cell.font.bold,
                color="FF000000",
            )


def main():
    wb = load_workbook(INVENTORY_FILE)
    ws = wb.active

    # Remove CF rules on L column — static fills are the source of truth
    to_remove = [r for r in ws.conditional_formatting._cf_rules
                 if "L" in str(r)]
    for r in to_remove:
        del ws.conditional_formatting._cf_rules[r]

    apply_fills(ws)
    fit_columns(wb[SALES_TAB])
    fit_columns(wb[WAREHOUSE_TAB])
    wb.save(INVENTORY_FILE)

    counts = {}
    for r in range(DATA_START, ws.max_row + 1):
        val = ws.cell(row=r, column=C_STATUS).value
        key = ("Allocated" if val and str(val).startswith("Allocated")
               else str(val) if val else "(none)")
        counts[key] = counts.get(key, 0) + 1

    print("Status fills applied:")
    for label, count in counts.items():
        if count:
            print(f"  {label:<16} {count:>3} row(s)")
    print(f"\nSaved → {INVENTORY_FILE}")


if __name__ == "__main__":
    main()
