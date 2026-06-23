import argparse
import csv
from openpyxl import load_workbook
from apply_status_fills import apply_fills

# ── CONFIG ────────────────────────────────────────────────────────────────────
MANIFEST_FILE  = "container_manifest.csv"
INVENTORY_FILE = "inventory_master.xlsx"

# Columns that together identify a unique product variant.
# Must match the column headers in the inventory sheet (row 2).
# "Top Shape" and "Hardware" were removed — the new template doesn't carry them.
VARIANT_KEYS = ["Design Name", "Size", "Finish", "Swing", "Glass Type"]

IN_STOCK_COL = "In-Stock QTY"
PRESALE_COL  = "Presale QTY"   # appears once per container group — resolved via group row
SERIAL_COL   = "Serial Number"  # optional: write serials here if the column exists

# Row layout of the new template (see create_inventory_master.py)
GROUP_ROW  = 1   # container labels: CNT-001, CNT-002, CNT-003
HEADER_ROW = 2   # column names
DATA_START = 3   # first product summary row


# ── LOAD MANIFEST ─────────────────────────────────────────────────────────────
def load_manifest(filename):
    with open(filename, newline="") as f:
        reader = csv.DictReader(f)
        return list(reader)


def variant_key(row, keys=VARIANT_KEYS):
    return tuple(row.get(k, "").strip() for k in keys)


# ── HEADER / COLUMN RESOLUTION ────────────────────────────────────────────────
def build_header_map(ws):
    """Map unique column headers (row 2) to their column numbers.

    Duplicate headers (e.g. "Presale QTY" × 3) are skipped after the first
    occurrence — use get_container_presale_col() to resolve those.
    """
    header_to_col = {}
    for cell in ws[HEADER_ROW]:
        if cell.value and cell.value not in header_to_col:
            header_to_col[cell.value] = cell.column
    return header_to_col


def get_container_presale_col(ws, container_id):
    """Return the column number of the Presale QTY cell that belongs to container_id.

    Finds the container label in the group row (row 1), then confirms the cell
    directly below it in the header row is PRESALE_COL.
    """
    for cell in ws[GROUP_ROW]:
        if cell.value == container_id:
            presale_header = ws.cell(row=HEADER_ROW, column=cell.column).value
            if presale_header != PRESALE_COL:
                raise ValueError(
                    f"Column {cell.column} under {container_id!r} has header "
                    f"{presale_header!r}, expected {PRESALE_COL!r}"
                )
            return cell.column
    raise ValueError(
        f"Container {container_id!r} not found in row {GROUP_ROW}. "
        f"Available containers: {[c.value for c in ws[GROUP_ROW] if c.value]}"
    )


# ── LOCATE PRODUCT ROWS IN INVENTORY SHEET ────────────────────────────────────
def find_product_rows(ws, header_to_col):
    """Map each product variant key to its summary row number.

    Scans from DATA_START so the two header rows are never matched.
    """
    rows_by_variant = {}
    for row_num in range(DATA_START, ws.max_row + 1):
        key = tuple(
            str(ws.cell(row=row_num, column=header_to_col[k]).value or "").strip()
            for k in VARIANT_KEYS
        )
        if any(key):
            rows_by_variant[key] = row_num
    return rows_by_variant


# ── MAIN UPDATE LOGIC ─────────────────────────────────────────────────────────
def update_inventory(manifest_items, ws, container_id):
    header_to_col = build_header_map(ws)

    for required in VARIANT_KEYS + [IN_STOCK_COL]:
        if required not in header_to_col:
            raise ValueError(f"Inventory sheet is missing expected column: {required!r}")

    presale_col = get_container_presale_col(ws, container_id)
    serial_col  = header_to_col.get(SERIAL_COL)  # None if column doesn't exist

    if serial_col is None:
        print(f"  Note: no '{SERIAL_COL}' column found — serial rows will be "
              f"inserted but serials won't be written to any cell.\n")

    rows_by_variant = find_product_rows(ws, header_to_col)

    # Resolve matches up front, then apply bottom-to-top so inserting serial
    # rows doesn't shift any product row we still need to visit.
    matches   = []
    unmatched = []
    for item in manifest_items:
        key = variant_key(item)
        if key in rows_by_variant:
            matches.append((rows_by_variant[key], item))
        else:
            unmatched.append(item)

    matches.sort(key=lambda m: m[0], reverse=True)

    for row_num, item in matches:
        qty = int(item["Quantity"])
        presale_cell  = ws.cell(row=row_num, column=presale_col)
        in_stock_cell = ws.cell(row=row_num, column=header_to_col[IN_STOCK_COL])

        presale_qty  = int(presale_cell.value or 0)
        in_stock_qty = int(in_stock_cell.value or 0)

        if qty > presale_qty:
            print(
                f"  WARNING: {item['Design Name']} ({item['Size']}, {item['Swing']}) "
                f"moving {qty} but only {presale_qty} in presale for {container_id}."
            )

        presale_cell.value  = presale_qty  - qty
        in_stock_cell.value = in_stock_qty + qty

        serials = [s for s in item.get("Serial Numbers", "").split(";") if s]
        if serials:
            ws.insert_rows(row_num + 1, amount=len(serials))
            if serial_col:
                for offset, serial in enumerate(serials):
                    ws.cell(row=row_num + 1 + offset, column=serial_col, value=serial)

        print(
            f"  Updated {item['Design Name']} ({item['Size']}, {item['Swing']}): "
            f"presale {presale_qty} -> {presale_qty - qty}, "
            f"in-stock {in_stock_qty} -> {in_stock_qty + qty}, "
            f"{len(serials)} serial(s) added"
        )

    return unmatched


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Move presale stock to in-stock when a container arrives."
    )
    parser.add_argument(
        "--container",
        required=True,
        help="Container ID whose presale column to decrement (e.g. CNT-001)",
    )
    parser.add_argument("--manifest",   default=MANIFEST_FILE,  help="Manifest CSV path")
    parser.add_argument("--inventory",  default=INVENTORY_FILE, help="Inventory XLSX path")
    args = parser.parse_args()

    print(f"Loading manifest from {args.manifest}...")
    manifest_items = load_manifest(args.manifest)
    print(f"Found {len(manifest_items)} line item(s).\n")

    print(f"Opening inventory workbook: {args.inventory}")
    wb = load_workbook(args.inventory)
    ws = wb.active

    print(f"Processing container: {args.container}\n")
    unmatched = update_inventory(manifest_items, ws, args.container)

    apply_fills(ws)
    wb.save(args.inventory)
    print(f"\nSaved updates to {args.inventory}")

    if unmatched:
        print(f"\n{len(unmatched)} manifest line item(s) had no matching inventory row:")
        for item in unmatched:
            key = " | ".join(item.get(k, "") for k in VARIANT_KEYS)
            print(f"  - {key}")


if __name__ == "__main__":
    main()
