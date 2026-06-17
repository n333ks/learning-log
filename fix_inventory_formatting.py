#!/usr/bin/env python3
"""
Fix four formatting issues in inventory_master.xlsx.

1. Black font on all body cells (rows 2+).  Strips any coloured-font bleed
   that occurred because the old STATUS CF applied to the whole row (A:L).

2. Variance column (J): add a CellIsRule — red fill + red font when the value
   is negative.  No rule for >=0 needed; the base black font covers it.
   Previous file had no Variance CF at all; this adds it.

3. Status column (L) CF only: coloured fill, explicit black font.
   Applies to col L only so it never overrides Variance or other columns.
   Uses full 8-char ARGB ("FF…") so Excel renders colours as fully opaque.

4. Design-block border: medium outer border wrapping every consecutive group
   of rows sharing the same Design Name.  The function is fully idempotent —
   it resets all data borders to thin first, then redraws — so it produces
   correct results after any add / remove / reorder without leaving stale
   thick lines behind.
"""

from openpyxl import load_workbook
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

INVENTORY_FILE = "inventory_master.xlsx"
HEADER_ROW     = 1
DATA_START     = 2
TOTAL_COLS     = 12          # A (Design Name) … L (Status)
VARIANCE_COL   = get_column_letter(10)   # J
STATUS_COL     = get_column_letter(12)   # L

# All colours use full 8-char ARGB (FF prefix = fully opaque).
STATUS_CF = [
    # (excel_comparison_value, CellIsRule_operator, fill_ARGB)
    ('"In Stock"',      "equal", "FFC6EFCE"),   # green
    ('"Pre-Sale"',      "equal", "FFFFEB9C"),   # yellow
    ('"In Production"', "equal", "FFFFC7CE"),   # red (light)
]
ALLOCATED_FILL = "FFBDD7EE"   # blue  (prefix match — FormulaRule)
VARIANCE_NEG_FILL = "FFFFCCCC"
VARIANCE_NEG_FONT = "FFC00000"  # dark red


# ── Dynamic design-block border ───────────────────────────────────────────────
def apply_design_block_borders(ws, data_start=DATA_START, total_cols=TOTAL_COLS):
    """
    Wrap each group of consecutive rows sharing the same Design Name (col A)
    with a medium outer border.  All other cell borders are reset to thin.

    Algorithm
    ---------
    1. Reset every data cell to a uniform thin border.  This clears any stale
       thick borders left by a previous call or by deleted rows.
    2. Scan col A from data_start downward; collect (first_row, last_row, name)
       spans — one span per run of identical Design Name values.
    3. For each span, redraw the four outer edges as medium borders while
       keeping inner horizontal and vertical borders thin.

    Returns the list of (first_row, last_row, design_name) tuples.
    """
    thin  = Side(border_style="thin",   color="BFBFBF")
    thick = Side(border_style="medium", color="404040")

    last_row = ws.max_row
    if last_row < data_start:
        return []

    # Step 1 — reset all data borders to thin
    for r in range(data_start, last_row + 1):
        for c in range(1, total_cols + 1):
            ws.cell(row=r, column=c).border = Border(
                top=thin, bottom=thin, left=thin, right=thin
            )

    # Step 2 — locate block boundaries
    blocks = []
    current = None
    block_start = None
    for r in range(data_start, last_row + 1):
        name = ws.cell(row=r, column=1).value
        if name != current:
            if current is not None:
                blocks.append((block_start, r - 1, current))
            current = name
            block_start = r
    if current is not None:
        blocks.append((block_start, last_row, current))

    # Step 3 — draw thick outer border for each block
    for first_r, last_r, _name in blocks:
        for r in range(first_r, last_r + 1):
            for c in range(1, total_cols + 1):
                ws.cell(row=r, column=c).border = Border(
                    top    = thick if r == first_r    else thin,
                    bottom = thick if r == last_r     else thin,
                    left   = thick if c == 1          else thin,
                    right  = thick if c == total_cols else thin,
                )

    return blocks


# ── Core formatting fix ───────────────────────────────────────────────────────
def fix_formatting(ws):
    """Apply all four formatting fixes to ws in place."""
    last_row = ws.max_row

    # ── Fix 1: black font on every body cell ─────────────────────────────────
    for r in range(DATA_START, last_row + 1):
        for c in range(1, TOTAL_COLS + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(
                name="Arial", size=10,
                bold=cell.font.bold,   # preserve any bold already set
                color="FF000000",
            )

    # ── Fix 2 & 3: replace all CF rules ──────────────────────────────────────
    ws.conditional_formatting = ConditionalFormattingList()

    var_range    = f"{VARIANCE_COL}{DATA_START}:{VARIANCE_COL}{last_row}"
    status_range = f"{STATUS_COL}{DATA_START}:{STATUS_COL}{last_row}"

    # Fix 2 — Variance: negative → light-red fill + dark-red font
    ws.conditional_formatting.add(
        var_range,
        CellIsRule(
            operator="lessThan",
            formula=["0"],
            fill=PatternFill("solid", fgColor=VARIANCE_NEG_FILL),
            font=Font(name="Arial", size=10, color=VARIANCE_NEG_FONT),
        ),
    )

    # Fix 3a — Status exact matches (CellIsRule, black font)
    for excel_value, operator, fill_argb in STATUS_CF:
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(
                operator=operator,
                formula=[excel_value],
                fill=PatternFill("solid", fgColor=fill_argb),
                font=Font(name="Arial", size=10, color="FF000000"),
            ),
        )

    # Fix 3b — Allocated prefix (FormulaRule; $L ref shifts by row automatically)
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=[f'=LEFT($L{DATA_START},9)="Allocated"'],
            fill=PatternFill("solid", fgColor=ALLOCATED_FILL),
            font=Font(name="Arial", size=10, color="FF000000"),
        ),
    )

    # ── Fix 4: design-block borders ──────────────────────────────────────────
    blocks = apply_design_block_borders(ws)

    return blocks


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print(f"Opening {INVENTORY_FILE} ...")
    wb = load_workbook(INVENTORY_FILE)
    ws = wb.active

    blocks = fix_formatting(ws)
    wb.save(INVENTORY_FILE)

    last_row = ws.max_row
    print(f"\n  Fix 1 — Black font applied to {last_row - HEADER_ROW} "
          f"body row(s) × {TOTAL_COLS} columns.")
    print(f"  Fix 2 — Variance CF on {VARIANCE_COL}{DATA_START}:{VARIANCE_COL}{last_row}"
          f"  (red fill + red font when < 0).")
    print(f"  Fix 3 — Status CF on {STATUS_COL}{DATA_START}:{STATUS_COL}{last_row}"
          f"  (coloured fill, black font, col L only).")
    print(f"  Fix 4 — Design-block borders drawn ({len(blocks)} block(s)):")
    for first_r, last_r, name in blocks:
        count = last_r - first_r + 1
        print(f"           {name:<14}  rows {first_r}–{last_r}  "
              f"({count} row{'s' if count != 1 else ''})")

    print(f"\n  Saved → {INVENTORY_FILE}")


if __name__ == "__main__":
    main()
